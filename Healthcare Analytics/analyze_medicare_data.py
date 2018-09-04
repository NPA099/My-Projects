
# coding: utf-8

# In[1]:

#Importing packages
import requests
import os
import zipfile
import openpyxl
import sqlite3
import glob
import csv
import numpy as np
from numpy import array


# In[2]:

#url for the accessing the medicare data
url = "https://data.medicare.gov/views/bg9k-emty/files/0a9879e0-3312-4719-a1db-39fd114890f1?content_type=application%2Fzip%3B%20charset%3Dbinary&filename=Hospital_Revised_Flatfiles.zip"


# In[3]:

#Accessing the url
r = requests.get(url)


# In[4]:

#Defining the staging directory name
staging_dir_name = "staging"


# In[5]:

staging_dir_name = "staging"
os.mkdir(staging_dir_name)


# In[6]:

#Checking if it is directory
os.path.isdir(staging_dir_name)


# In[7]:

#Zip file name and path
zip_file_name = os.path.join(staging_dir_name,"test.zip")


# In[8]:

#Checking if the zip file is created in the respective directory
zip_file_name


# In[9]:

#"wb" is used to write in binary format
zf = open(zip_file_name,"wb") 


# In[10]:

#Writing the zip file content
zf.write(r.content)


# In[11]:

#Closing the file
zf.close()


# In[12]:

#'r' is used to read the file.......... extractall will unzip the contents
z = zipfile.ZipFile(zip_file_name,'r')
#Unzipping the test.zip file
z.extractall(staging_dir_name)
#Closing the zip file
z.close()


# In[13]:

#Function to convert cp1252 to utf-8 format
def cp_to_utf (glob_dir):
    for file_name in glob.glob(glob_dir):
        print (file_name[8:])
        fn = os.path.join(staging_dir_name, file_name[8:])
        #print ("*********************")
        in_fp = open(fn,"rt",encoding = 'cp1252')
        input_data = in_fp.read()
        in_fp.close()
        ofn = os.path.join(staging_dir_name, file_name[8:])
        out_fp = open(ofn,"wt",encoding = 'utf-8')
        for c in input_data:
            if c != '\0':
                out_fp.write(c)
        out_fp.close()
        #print("$$$$$$$$$$$$$$$$$$$$$$$$$")
    #print("@@@@@@@@@@@@@@@@@@@@@@@@@@")


# In[14]:

#Converting the file name to table name as per the requirement
def tbl_name (glob_dir): 
    #print(file_name)
    #print(file_name[8:-4])
    File_Name = file_name[8:-4]
    Table_Name = File_Name
    Table_Name = glob_dir[8:-4]
    print(Table_Name)
    Table_Name = Table_Name.lower()
    Table_Name = Table_Name.replace(' ', '_')
    Table_Name = Table_Name.replace('-', '_')
    Table_Name = Table_Name.replace('%', 'pct')
    Table_Name = Table_Name.replace('/', '_')
    if(Table_Name[0].islower() == False):
        T = 't_'+Table_Name
        #print(T)
        return T;
        #print("TTTTTTTTTTTTTTTTTTTTTTTTTTTTT")
    else:
        T = Table_Name
        #print(T)
        #print("TTTTTTTTTTTTTTTTTTTTTTTTTTTTT")
        return T;


# In[15]:

#Converting the column name as per the requirement
def col_name (C): 
    #print(C)
    Column_Name = C
    #print(Column_Name)
    Column_Name = Column_Name.lower()
    Column_Name = Column_Name.replace(' ', '_')
    Column_Name = Column_Name.replace('-', '_')
    Column_Name = Column_Name.replace('%', 'pct')
    Column_Name = Column_Name.replace('/', '_')
    if(Column_Name[0].islower() == False):
        C = ['c_'+Column_Name]
        #print(C)
        return C;
        #print("CCCCCCCCCCCCCCCCCCCCCCCCCCCCC")
    else:
        C = [Column_Name]
        #print(C)
        #print("CCCCCCCCCCCCCCCCCCCCCCCCCCCCC")
        return C;


# In[16]:

#Defining the path of the file and calling function for UTF-8 encoding conversion
glob_dir = os.path.join(staging_dir_name,"*.csv")
cp_to_utf(glob_dir)


# In[17]:

#Creating tables in SQL lite and inserting the data
for file_name in glob.glob(glob_dir):
    if file_name[8:] != "FY2015_Percent_Change_in_Medicare_Payments.csv":    
        with open(file_name, 'r', encoding = 'UTF-8') as file:
            reader = csv.reader(file)
            Col_Name = next(reader)
            #print (file_name)
            #print (Col_Name)
            #print ("$$$$$$$$$$$$$$$$$$$$$$$$$$")
            Tab = tbl_name(file_name)
            Col = []
            for C in Col_Name:
                Col = Col+col_name(C)
            print (Col)
            print ("&&&&&&&&&&&&&&&&&&&&&&&&&")
            #Defining the connection to db
            conn = sqlite3.connect("medicare_hospital_compare.db")
            #Defining the table
            DDL = "create table if not exists " + Tab + " ("
            for C_Name in Col[:-1]:
                DDL = DDL+C_Name+" text, "
            DDL = DDL+Col[-1]+" text)"
            c1 = conn.cursor()
            #Executing query to create stage  tables
            c1.execute(DDL)
            #data = []
            #data_lst = []
            data_lst_lst = [row for row in reader if row != [' ']]
            data_lst_tup = [tuple(l) for  l in data_lst_lst]
            DML_I = "Insert into " + Tab + " values ("
            i = 1
            while (i<len(Col)):
                DML_I = DML_I + "?,"
                i += 1
            DML_I = DML_I + "?)"
            print (DML_I)
            c2 = conn.cursor()
            c2.executemany(DML_I, data_lst_tup)
            #Commiting to the db
            conn.commit()
            print (DDL)
        print ("%%%%%%%%%%%%%%%%%%%%%%%%%%%%")
    else:
        continue


# In[18]:

#Hospital ranking and focus states file location in web
k_url = "http://kevincrook.com/utd/hospital_ranking_focus_states.xlsx"


# In[19]:

#Accessing the url
r1 = requests.get(k_url)


# In[20]:

#Creating the excel file
xf = open("hospital_ranking_focus_states.xlsx","wb") 


# In[21]:

#Writing the data in file
xf.write(r1.content)


# In[22]:

#Closing the excel file
xf.close()


# In[23]:

#Creating hospital national ranking table in sqlite
c3 = conn.cursor()
DDL_hnr = """
Create table if not exists hospital_national_ranking (
provider_id text, 
ranking number
)
"""
c3.execute(DDL_hnr)
#Commiting to the db
conn.commit()


# In[24]:

#Inserting data into hospital_national_ranking table in sqlite
c4 = conn.cursor()
wb = openpyxl.load_workbook("hospital_ranking_focus_states.xlsx")
sheet=wb.get_sheet_by_name("Hospital National Ranking")
#List of columns
Col_tup = []
for a, rows in enumerate(sheet):
    Col_tup_1 = []
    if a == 0:
        continue
    for row in rows:
        if str(row.value).strip() != 'None':
            Col_tup_1.append(str(row.value).strip())
    Col_tup.append(Col_tup_1)
Col_lst_tup = [tuple(l) for l in Col_tup]
DDL_I1 = "insert into hospital_national_ranking values (?,?)"
c4.executemany(DDL_I1, tuple(Col_lst_tup))
#Commiting to the db
conn.commit()


# In[25]:

#Dictionery containing the focus states
Focus_states_dct={'California':'CA','Florida':'FL','Georgia':'GA','Illinois':'IL','Kansas':'KS','Michigan':'MI','New York':'NY','Ohio':'OH','Pennsylvania':'PA','Texas':'TX'}


# In[26]:

#Creating hospital ranking excel sheet
sheet2=wb.get_sheet_by_name("Focus States")
s=2
States_lst = []
while sheet2.cell(row=s,column=1).value!=None:
    State=[sheet2.cell(row=s, column=1).value]
    States_lst=States_lst+State
    s+=1
States_lst.insert(0, 'Nationwide')
#Importing workbook package from xlsxwriter
from xlsxwriter.workbook import Workbook
wrkbk = Workbook('hospital_ranking.xlsx')
#Initializing sheet number
c5 = conn.cursor()
Sh = 0
#Iterating the loop to create the sheets and write the data in respective sheets
while Sh<11:
    wrksheet = wrkbk.add_worksheet(States_lst[Sh])
    if States_lst[Sh] == 'Nationwide':
        #Defining the query to get the hospital national ranking data nation wide
        Sel = c5.execute("""Select hospital_national_ranking.provider_id, 
        hospital_name, city, state, county_name 
        from hospital_national_ranking join hospital_general_information
        on hospital_national_ranking.provider_id = hospital_general_information.provider_id
        order by ranking limit 100""")
    else:
        l=[Focus_states_dct[States_lst[Sh]]]
        #Defining the query to get the hospital national ranking data for the focus states
        Sel=c5.execute("""Select hospital_national_ranking.provider_id, 
        hospital_name, city, state, county_name 
        from hospital_national_ranking join hospital_general_information
        on hospital_national_ranking.provider_id = hospital_general_information.provider_id
        where state in ('"""+','.join(map(str, l)) + "') order by ranking limit 100")
    Headers = ["Provider ID", "Hospital Name", "City", "State", "County"]
    b=0
    while b<5:
        wrksheet.write(0, b, Headers[b])
        b+=1
    a=0
    for a, row in enumerate(Sel):
        #Writing the data to the hospital ranking sheet
        wrksheet.write(a+1, 0, row[0])
        wrksheet.write(a+1, 1, row[1])
        wrksheet.write(a+1, 2, row[2])
        wrksheet.write(a+1, 3, row[3])
        wrksheet.write(a+1, 4, row[4])
    Sh += 1
wrkbk.close()


# In[27]:

#Ignoring the non numeric values from timely_and_effective_care___hospital for measures excel sheet
c6 = conn.cursor()
#Defining query to get distinct measure id and measure name
Sel=c6.execute("""select distinct measure_id, measure_name 
from timely_and_effective_care___hospital 
where score not in 
('Not Available', 
'Very High (60,000+ patients annually)', 
'High (40,000 - 59,999 patients annually)', 
'Medium (20,000 - 39,999 patients annually)', 
'Low (0 - 19,999 patients annually)') 
order by measure_id""")

Measure_id=[]
Measure_name=[]
for row in Sel: 
    Measure_id.append(row[0])
    Measure_name.append(row[1])
    
Measure_map = dict(zip(Measure_id,Measure_name))

wrkbk = Workbook('measures_statistics.xlsx')
Sh = 0
c7=conn.cursor()
#Iterating the loop to create the sheets and write the data in respective sheets
while Sh<11:
    wrksheet = wrkbk.add_worksheet(States_lst[Sh])
    if States_lst[Sh] == 'Nationwide':
        #Defining query to get measure id, measure name and score
        Sel = c7.execute("""select measure_id, measure_name, score 
        from timely_and_effective_care___hospital 
        where score not in 
        ('Not Available', 
        'Very High (60,000+ patients annually)', 
        'High (40,000 - 59,999 patients annually)', 
        'Medium (20,000 - 39,999 patients annually)', 
        'Low (0 - 19,999 patients annually)') 
        order by measure_id""")
    else:
        l=[Focus_states_dct[States_lst[Sh]]]
        #Defining query to get measure id, measure name and score
        Sel = c7.execute("""select measure_id, measure_name, score 
        from timely_and_effective_care___hospital 
        where score not in 
        ('Not Available', 
        'Very High (60,000+ patients annually)', 
        'High (40,000 - 59,999 patients annually)', 
        'Medium (20,000 - 39,999 patients annually)', 
        'Low (0 - 19,999 patients annually)') 
        and state in ('"""+ ','.join(map(str,l))+"') order by measure_id") 
    
    Measure_values = []
    each_lst = []
    for row in Sel: 
        if row[2].isdigit():
            Measure_values.append(row[0])
    Measure_lst = np.unique(Measure_values).tolist()
    Headers=["Measure ID", "Measure Name", "Minimum", "Maximum", "Average", "Standard Deviation"]
    
    b=0
    while b<6:
        wrksheet.write(0, b, Headers[b])
        b += 1
    each_lst = []
    a=0
    while a<len(Measure_lst):
        if States_lst[Sh] == 'Nationwide':
            #Defining query to get measure id, measure name and score
            Sel = c7.execute("""select measure_id, measure_name, score 
            from timely_and_effective_care___hospital 
            where score not in 
            ('Not Available', 
            'Very High (60,000+ patients annually)', 
            'High (40,000 - 59,999 patients annually)', 
            'Medium (20,000 - 39,999 patients annually)', 
            'Low (0 - 19,999 patients annually)') 
            order by measure_id""")
        else: 
            l=[Focus_states_dct[States_lst[Sh]]] 
            #Defining query to get measure id, measure name and score
            Sel = c7.execute("""select measure_id, measure_name, score 
            from timely_and_effective_care___hospital 
            where score not in 
            ('Not Available', 
            'Very High (60,000+ patients annually)', 
            'High (40,000 - 59,999 patients annually)', 
            'Medium (20,000 - 39,999 patients annually)', 
            'Low (0 - 19,999 patients annually)') 
            and state in ('"""+ ','.join(map(str,l))+"') order by measure_id") 
        each_lst=[]
        for row in Sel:
            if row[0]==(Measure_lst[a]):
                each_lst.append(row[2])
        each_lst = list(map(int, each_lst))
        data = np.asarray(each_lst)
        #Writing the measures to the measure statistics excel sheet
        wrksheet.write(a+1, 0, Measure_lst[a]) 
        wrksheet.write(a+1, 1, Measure_map[Measure_lst[a]]) 
        wrksheet.write(a+1, 2, data.min()) 
        wrksheet.write(a+1, 3, data.max()) 
        wrksheet.write(a+1, 4, data.mean()) 
        wrksheet.write(a+1, 5, data.std()) 
        a+=1
    Sh+=1
wrkbk.close()

